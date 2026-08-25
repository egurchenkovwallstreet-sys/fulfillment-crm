"""Импорт остатков из Excel (формат WB): баркод + количество, прибавление к CRM и WB."""
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from django.db import transaction

from apps.integrations.models import AuditLog
from apps.sellers.models import Seller
from apps.warehouse.models import Product, ProductWarehouseStock, StockOperation
from apps.warehouse.services.catalog_fetch import CatalogError, build_seller_catalog_index
from apps.warehouse.services.cells import create_cell_with_next_number, refresh_cell_occupied
from apps.warehouse.services.wb_stocks import (
  WBStockError,
  fetch_wb_stock_for_barcode,
  fetch_wb_stocks_for_warehouses,
  get_seller_warehouse,
  increment_product_warehouse_stock,
  push_wb_stock_increment,
)

try:
  from openpyxl import load_workbook
except ImportError:  # pragma: no cover
  load_workbook = None


class StockFileImportError(Exception):
  pass


BARCODE_HEADERS = {
  "баркод", "barcode", "sku", "штрихкод", "штрих-код", "штрих код",
}
QTY_HEADERS = {
  "количество", "кол-во", "кол во", "колво", "amount", "qty", "остаток", "quantity",
}


@dataclass
class ParsedStockRow:
  barcode: str
  add_quantity: int


@dataclass
class StockImportPreviewRow:
  barcode: str
  add_quantity: int
  status: str
  title: str
  crm_before: int
  crm_after: int
  wb_before: int
  wb_after: int
  will_create: bool
  cell_number: str
  message: str


def _normalize_header(value) -> str:
  text = str(value or "").strip().lower()
  text = text.replace("\xa0", " ")
  return re.sub(r"\s+", " ", text)


def _parse_quantity(value) -> int | None:
  if value is None or value == "":
    return None
  if isinstance(value, (int, float)):
    qty = int(value)
    return qty if qty > 0 else None
  text = str(value).strip().replace(",", ".")
  if not text:
    return None
  try:
    qty = int(float(text))
  except ValueError:
    return None
  return qty if qty > 0 else None


def parse_stock_excel(file_bytes: bytes) -> list[ParsedStockRow]:
  if load_workbook is None:
    raise StockFileImportError("На сервере не установлен openpyxl для чтения Excel")

  try:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
  except Exception as exc:
    raise StockFileImportError(f"Не удалось прочитать Excel: {exc}") from exc

  sheet = workbook.active
  rows = list(sheet.iter_rows(values_only=True))
  if not rows:
    raise StockFileImportError("Файл пустой")

  header_idx = None
  barcode_col = None
  qty_col = None

  for idx, row in enumerate(rows[:20]):
    headers = [_normalize_header(cell) for cell in row]
    bc_candidates = [i for i, h in enumerate(headers) if h in BARCODE_HEADERS]
    qty_candidates = [i for i, h in enumerate(headers) if h in QTY_HEADERS]
    if bc_candidates and qty_candidates:
      header_idx = idx
      barcode_col = bc_candidates[0]
      qty_col = qty_candidates[0]
      break

  if header_idx is None or barcode_col is None or qty_col is None:
    raise StockFileImportError(
      "Не найдены колонки «Баркод» и «Количество». Используйте выгрузку остатков WB."
    )

  aggregated: dict[str, int] = {}
  for row in rows[header_idx + 1:]:
    if not row:
      continue
    barcode = str(row[barcode_col] or "").strip()
    qty = _parse_quantity(row[qty_col] if qty_col < len(row) else None)
    if not barcode or qty is None:
      continue
    aggregated[barcode] = aggregated.get(barcode, 0) + qty

  if not aggregated:
    raise StockFileImportError("В файле нет строк с баркодом и количеством")

  return [
    ParsedStockRow(barcode=barcode, add_quantity=qty)
    for barcode, qty in sorted(aggregated.items())
  ]


def _get_crm_warehouse_qty(product: Product | None, warehouse) -> int:
  if not product:
    return 0
  pws = ProductWarehouseStock.objects.filter(
    product=product,
    seller_warehouse=warehouse,
  ).first()
  return int(pws.quantity) if pws else 0


def build_stock_import_preview(
  seller: Seller,
  *,
  warehouse_id: int,
  file_bytes: bytes,
) -> dict:
  warehouse = get_seller_warehouse(seller, warehouse_id)
  parsed_rows = parse_stock_excel(file_bytes)
  file_units = sum(row.add_quantity for row in parsed_rows)

  try:
    catalog_index = build_seller_catalog_index(seller)
  except CatalogError as exc:
    raise StockFileImportError(str(exc)) from exc

  crm_products = {
    p.barcode: p
    for p in Product.objects.filter(seller=seller).select_related("cell")
  }

  known_barcodes = [row.barcode for row in parsed_rows if row.barcode in catalog_index]
  try:
    wb_stock_map = fetch_wb_stocks_for_warehouses(seller, [warehouse], known_barcodes)
  except WBStockError as exc:
    raise StockFileImportError(str(exc)) from exc

  preview_rows: list[StockImportPreviewRow] = []
  skipped_unknown_details: list[dict] = []

  for row in parsed_rows:
    catalog_item = catalog_index.get(row.barcode)
    if not catalog_item:
      skipped_unknown_details.append({
        "barcode": row.barcode,
        "add_quantity": row.add_quantity,
      })
      continue

    product = crm_products.get(row.barcode)
    crm_before = product.quantity if product else 0
    wb_before = int((wb_stock_map.get(row.barcode) or {}).get("total") or 0)

    preview_rows.append(
      StockImportPreviewRow(
        barcode=row.barcode,
        add_quantity=row.add_quantity,
        status="ok",
        title=catalog_item.title,
        crm_before=crm_before,
        crm_after=crm_before + row.add_quantity,
        wb_before=wb_before,
        wb_after=wb_before + row.add_quantity,
        will_create=product is None,
        cell_number=product.cell.number if product else "",
        message="",
      ),
    )

  return {
    "warehouse": {
      "id": warehouse.id,
      "wb_warehouse_id": warehouse.wb_warehouse_id,
      "name": warehouse.name,
    },
    "rows": [_serialize_preview_row(row) for row in preview_rows],
    "skipped_unknown": [item["barcode"] for item in skipped_unknown_details],
    "skipped_unknown_details": skipped_unknown_details,
    "totals": {
      "file_barcodes": len(parsed_rows),
      "file_units": file_units,
      "to_apply": len(preview_rows),
      "skipped_unknown": len(skipped_unknown_details),
      "skipped_units": sum(item["add_quantity"] for item in skipped_unknown_details),
      "new_products": sum(1 for row in preview_rows if row.will_create),
      "add_units": sum(row.add_quantity for row in preview_rows),
    },
  }


def _serialize_preview_row(row: StockImportPreviewRow) -> dict:
  return {
    "barcode": row.barcode,
    "add_quantity": row.add_quantity,
    "status": row.status,
    "title": row.title,
    "crm_before": row.crm_before,
    "crm_after": row.crm_after,
    "wb_before": row.wb_before,
    "wb_after": row.wb_after,
    "will_create": row.will_create,
    "cell_number": row.cell_number,
    "message": row.message,
  }


def _serialize_mismatch(item: dict) -> dict:
  return {
    "barcode": item["barcode"],
    "add_quantity": item["add_quantity"],
    "crm_before": item["crm_before"],
    "crm_expected": item["crm_expected"],
    "crm_actual": item["crm_actual"],
    "wb_before": item["wb_before"],
    "wb_expected": item["wb_expected"],
    "wb_actual": item["wb_actual"],
    "error": item["error"],
    "stage": item["stage"],
  }


@transaction.atomic
def apply_stock_import(
  seller: Seller,
  *,
  warehouse_id: int,
  rows: list[dict],
  user=None,
) -> dict:
  warehouse = get_seller_warehouse(seller, warehouse_id)
  if not rows:
    raise StockFileImportError("Нет строк для применения")

  try:
    catalog_index = build_seller_catalog_index(seller)
  except CatalogError as exc:
    raise StockFileImportError(str(exc)) from exc

  aggregated: dict[str, int] = {}
  for row in rows:
    barcode = str(row.get("barcode") or "").strip()
    try:
      qty = int(row.get("add_quantity") or 0)
    except (TypeError, ValueError):
      qty = 0
    if not barcode or qty <= 0:
      continue
    aggregated[barcode] = aggregated.get(barcode, 0) + qty

  if not aggregated:
    raise StockFileImportError("Нет корректных строк для применения")

  applied = 0
  created_products = 0
  verified = 0
  skipped_unknown_details: list[dict] = []
  mismatches: list[dict] = []

  was_crm_units = 0
  was_wb_units = 0
  added_units = 0
  result_crm_units = 0
  result_wb_units = 0

  for barcode, add_qty in aggregated.items():
    catalog_item = catalog_index.get(barcode)
    if not catalog_item:
      skipped_unknown_details.append({"barcode": barcode, "add_quantity": add_qty})
      mismatches.append({
        "barcode": barcode,
        "add_quantity": add_qty,
        "crm_before": 0,
        "crm_expected": 0,
        "crm_actual": 0,
        "wb_before": 0,
        "wb_expected": 0,
        "wb_actual": 0,
        "error": "Баркод не найден в каталоге WB селлера",
        "stage": "catalog",
      })
      continue

    product = (
      Product.objects.select_for_update()
      .filter(seller=seller, barcode=barcode)
      .select_related("cell")
      .first()
    )
    crm_before = product.quantity if product else 0
    wb_before = fetch_wb_stock_for_barcode(seller, warehouse, barcode)
    crm_wh_before = _get_crm_warehouse_qty(product, warehouse)
    crm_expected = crm_before + add_qty
    wb_expected = wb_before + add_qty
    crm_wh_expected = crm_wh_before + add_qty

    was_crm_units += crm_before
    was_wb_units += wb_before

    savepoint = transaction.savepoint()
    created_here = False
    try:
      if product:
        product.quantity += add_qty
        product.save(update_fields=["quantity", "updated_at"])
        increment_product_warehouse_stock(product, warehouse, add_qty)
      else:
        cell = create_cell_with_next_number(seller)
        product = Product.objects.create(
          seller=seller,
          barcode=barcode,
          name=catalog_item.title,
          cell=cell,
          quantity=add_qty,
          requires_marking=catalog_item.requires_marking,
          wb_nm_id=catalog_item.wb_nm_id,
          vendor_code=catalog_item.vendor_code,
          tech_size=catalog_item.tech_size,
          wb_size=catalog_item.wb_size,
          photo_url=catalog_item.photo_url,
        )
        refresh_cell_occupied(cell)
        increment_product_warehouse_stock(product, warehouse, add_qty)
        created_here = True

      push_wb_stock_increment(seller, warehouse, barcode, add_qty)

      product.refresh_from_db()
      crm_actual = product.quantity
      crm_wh_actual = _get_crm_warehouse_qty(product, warehouse)
      wb_actual = fetch_wb_stock_for_barcode(seller, warehouse, barcode)

      crm_ok = crm_actual == crm_expected and crm_wh_actual == crm_wh_expected
      wb_ok = wb_actual == wb_expected

      if not crm_ok and not wb_ok:
        raise StockFileImportError(
          f"CRM: ожидалось {crm_expected}, получилось {crm_actual}; "
          f"WB: ожидалось {wb_expected}, получилось {wb_actual}",
        )
      if not crm_ok:
        raise StockFileImportError(
          f"CRM: ожидалось {crm_expected} (склад {crm_wh_expected}), "
          f"получилось {crm_actual} (склад {crm_wh_actual})",
        )
      if not wb_ok:
        raise StockFileImportError(
          f"WB: ожидалось {wb_expected}, получилось {wb_actual}",
        )

      StockOperation.objects.create(
        product=product,
        operation_type=StockOperation.OperationType.INTAKE,
        quantity=add_qty,
        performed_by=user,
        comment=(
          f"Импорт Excel +{add_qty} шт., склад WB "
          f"{warehouse.name or warehouse.wb_warehouse_id}"
        ),
      )
      transaction.savepoint_commit(savepoint)
      applied += 1
      verified += 1
      added_units += add_qty
      if created_here:
        created_products += 1
      result_crm_units += crm_actual
      result_wb_units += wb_actual
    except (WBStockError, StockFileImportError) as exc:
      transaction.savepoint_rollback(savepoint)
      if isinstance(exc, WBStockError):
        stage = "wb"
      elif "CRM" in str(exc):
        stage = "crm"
      elif "WB" in str(exc):
        stage = "wb"
      else:
        stage = "verify"

      mismatches.append({
        "barcode": barcode,
        "add_quantity": add_qty,
        "crm_before": crm_before,
        "crm_expected": crm_expected,
        "crm_actual": crm_before,
        "wb_before": wb_before,
        "wb_expected": wb_expected,
        "wb_actual": wb_before,
        "error": str(exc),
        "stage": stage,
      })
      result_crm_units += crm_before
      result_wb_units += wb_before

  file_barcodes = len(aggregated)
  file_units = sum(aggregated.values())

  all_ok = len(mismatches) == 0 and applied > 0

  summary = {
    "file_barcodes": file_barcodes,
    "file_units": file_units,
    "was_crm_units": was_crm_units,
    "was_wb_units": was_wb_units,
    "added_units": added_units,
    "expected_crm_units": was_crm_units + added_units,
    "expected_wb_units": was_wb_units + added_units,
    "result_crm_units": result_crm_units,
    "result_wb_units": result_wb_units,
    "applied_barcodes": applied,
    "verified_barcodes": verified,
    "failed_barcodes": len(mismatches),
  }

  AuditLog.objects.create(
    user=user,
    seller=seller,
    action_type=AuditLog.ActionType.INTAKE,
    message=(
      f"Импорт остатков Excel: {applied}/{file_barcodes} баркодов, "
      f"+{added_units} шт., сверка {'OK' if all_ok else 'ОШИБКИ'}"
    ),
    details={
      "warehouse_id": warehouse.id,
      "summary": summary,
      "applied": applied,
      "created_products": created_products,
      "skipped_unknown_details": skipped_unknown_details,
      "mismatches": mismatches,
      "all_ok": all_ok,
    },
  )

  if applied == 0 and mismatches:
    raise StockFileImportError(
      f"Не удалось применить ни одного баркода. Ошибок: {len(mismatches)}",
    )

  return {
    "ok": all_ok,
    "applied": applied,
    "created_products": created_products,
    "verified": verified,
    "skipped_unknown": [item["barcode"] for item in skipped_unknown_details],
    "skipped_unknown_details": skipped_unknown_details,
    "mismatches": [_serialize_mismatch(item) for item in mismatches],
    "summary": summary,
    "add_units": added_units,
  }
